from openpyxl import Workbook
wb = Workbook()

#grab the active worksheet
ws = wb.active

#Data can be assigned directly to cells
ws['A1'] = 42

#Rows can also be append
ws.append([1,2,3])

#python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

#save the file
wb.save("sample1.xlsx")